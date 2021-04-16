import os
import nltk
import string

from openpyxl import load_workbook
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
nltk.download('stopwords')
nltk.download('wordnet')

def processFile(f,name):
	print("Start Processing: %s" % name)

	file = open(f, "rb")
	readContents = file.read().decode('iso-8859-1').splitlines()
	word_list = []
	for line in readContents:
		lineSplit = line.split(" ")
		for word in lineSplit:
			if word.endswith("in'") == True:
				l = list(word)
				l[len(word)-1] = 'g'
				word = ''.join(l)
			word_list.append(word.lower())

	censor = stopwords.words('english')
	filtered_words = [word for word in word_list if word not in censor]
	result_str = ' '.join(filtered_words)

	translator = str.maketrans('', '', string.punctuation)
	result_str = result_str.translate(translator)

	result_str = result_str.lower()
	result_arr = result_str.split(" ")

	lemmaWorkbook = load_workbook('lemmas.xlsx')
	lemmer = WordNetLemmatizer()

	word_count = len(result_arr)

	lemmaList = lemmaWorkbook.worksheets[0]
	finalList = []
	prefix = ""
	preflag = False
	for ix, word in enumerate(result_arr):
		flag = False
		for i in range(1,lemmaList.max_row):
			if lemmaList.cell(row=i, column=1).value == word:
				finalList.append(lemmaList.cell(row=i, column=3).value)
				flag = True
		if flag == False:
			word = lemmer.lemmatize(word)
			finalList.append(word)
		word = finalList[ix]
		if (word == "in#") or (word == "un#") or (word == "dis#") or (word == "non#") or (word == "de#") or (word == "re#") or (word == "mis#"):
			prefix = word[:len(word)-1]
			preflag = True
		else:
			if preflag == True:
				finalList[ix] = prefix+word
				preflag = False
			else:
				finalList[ix] = word

	final_str = ' '.join(finalList)
	product = open('./processed_lyrics/%s' % name, "w")
	product.write(final_str)


rootDir = './input_lyrics'
destinationDir = './processed_lyrics'
counter = 0

ix = len(os.listdir(destinationDir))-1 #redo last file
if ix < 0:
	ix = 0
print(ix)
for dirName, subdirList, fileList in os.walk(rootDir):
	print('Found directory: %s' % dirName)
	total = len(fileList)
	while ix < total:
		file = rootDir + "/" + fileList[ix]
		processFile(file,fileList[ix])
		print("%s :: %d/%d || COMPLETE" % (fileList[ix], ix+1, total))
		ix += 1